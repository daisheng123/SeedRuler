# -*- coding: utf-8 -*-
"""
Created on Thu Oct 26 11:45:44 2023

@author: 1595


"""

import numpy as np
from scipy.linalg import eig
import openpyxl
from scipy.linalg import eigh

def FitEllip(X, Y, N):
    # 打开Excel文件
    workbook = openpyxl.load_workbook('D:\B_GENERAL\coding transform\project\matlabGUI\data3.xlsx')
       
    # 选择要读取的工作表
    worksheet = workbook['Sheet1']
    
    # 创建一个空数组用于存储数据
    X = []
    
    # 遍历工作表的每一行
    for row in worksheet.iter_rows(values_only=True):
        X.append(row)
    X=np.array(X).flatten()
    workbook = openpyxl.load_workbook('D:\B_GENERAL\coding transform\project\matlabGUI\data4.xlsx')
    
    # 选择要读取的工作表
    worksheet = workbook['Sheet1']
    
    # 创建一个空数组用于存储数据
    Y = []
    
    # 遍历工作表的每一行
    for row in worksheet.iter_rows(values_only=True):
        Y.append(row)
    Y=np.array(Y).flatten()
    
    N = 720
    mx = np.mean(X)
    my = np.mean(Y)
    sx = (np.max(X) - np.min(X)) / 2
    sy = (np.max(Y) - np.min(Y)) / 2
    x = (X - mx) / sx
    y = (Y - my) / sy

    D = np.column_stack((x*x, x*y, y*y, x, y, np.ones_like(x)))
    S = np.dot(D.T, D)
    # print("sx",sx)
    # print("sy",sy)
    # print("mx",mx)
    # print("my",my)
    #print("D",D)
    #print("D_",D.shape)
    #print("S",S)
    
    C = np.zeros((6, 6))
    C[0, 2] = -2
    C[1, 1] = 1
    C[2, 0] = -2

    workbook = openpyxl.load_workbook('D:\B_GENERAL\coding transform\project\matlabGUI\data5.xlsx')
       
    # 选择要读取的工作表
    worksheet = workbook['Sheet1']
    
    # 创建一个空数组用于存储数据
    S = []
    
    # 遍历工作表的每一行
    for row in worksheet.iter_rows(values_only=True):
        S.append(row)
    S=np.array(S)
    #print("S",S)
    
    #geval, gevec = eig(S, C)

    # # Find the negative eigenvalue
    # indices = np.where((geval < 0) & (~np.isinf(geval)))
    # NegR1 = indices[0]
    # NegC = indices[0]
    # #print(indices)
    # A = gevec[:, NegC]
    
    
    geval, gevec = eig(S, C)

    # 对特征向量进行排序
    sort_indices = np.argsort(geval)[::-1]
    geval = geval[sort_indices]
    gevec = gevec[:, sort_indices]
    neg_indices = np.where((geval < 0) & (~np.isinf(geval)))[0]

    # 提取对应于正特征值的特征向量
    A = gevec[:, neg_indices]


    print("gevec",gevec)
    print("geval",geval)
    print('A',A)
    a = np.zeros(6)

    a[0] = A[0] * sy * sy
    a[1] = A[1] * sx * sy
    a[2] = A[2] * sx * sx
    a[3] = -2 * A[0] * sy * sy * mx - A[1] * sx * sy * my + A[3] * sx * sy * sy
    a[4] = -A[1] * sx * sy * mx - 2 * A[2] * sx * sx * my + A[5] * sx * sx * sy
    a[5] = A[0] * sy * sy * mx * mx + A[1] * sx * sy * mx * my + A[2] * sx * sx * my * my - A[3] * sx * sy * sy * mx - A[4] * sx * sx * sy * my + A[5] * sx * sx * sy * sy
    a = a.reshape(-1, 1)
    
    theta = np.arctan2(np.real(a[1]), np.real(a[0] - a[2])) / 2
    #theta = np.squeeze(theta)
    ct = np.cos(theta)
    st = np.sin(theta)
    ap = a[0] * ct * ct + a[1] * ct * st + a[2] * st * st
    cp = a[0] * st * st - a[1] * ct * st + a[2] * ct * ct
    T = np.array([[a[0], a[1]/2], [a[1]/2, a[2]]])
    T = np.squeeze(T)  # 去除多余的维度
    print("a[0]",a[0])
    print("a[1]",a[1])
    print("a[2]",a[2])
    print("a[3]",a[3])
    print("a[4]",a[4])
    print("a[5]",a[5])
    # print("T",T)
    # print("ap",ap)
    # print("cp",cp)

    # if T.shape[0] != T.shape[1]:
    #     # 处理不是方阵的情况
    #     print("输入矩阵不是方阵")
    # else:
    #     # 计算逆矩阵
    #     inv_T = np.linalg.inv(T)
    #     print(inv_T)

    #t = -np.linalg.inv(2 * T) @ np.array([[a[3], a[4]]]).T
    #t = -np.linalg.inv(2*T) @ np.array([[a[3]], [a[4]]])
    #t = -np.linalg.inv(2*T) @ np.array([[a[3]], [a[4]]])
    t = -np.linalg.inv(2*T) @ np.reshape(np.array([a[3], a[4]]), (2, 1))

    cx = t[0]
    cy = t[1]
    val = t.T @ T @ t
    scale = 1 / (val - a[5])
    #scale = np.squeeze(scale)
    #print("cx",cx)
    #print("cy",cy)
    print("val",val)
    print("scale",scale)
    
    r1 = 1 / np.sqrt(scale * ap)
    r2 = 1 / np.sqrt(scale * cp)
    
   
    v = np.array([r1, r2, cx, cy, theta], dtype=object).reshape(-1, 1)
    
    if r1 < r2:
        v = np.array([r2, r1, cx, cy, theta + np.pi/2]).reshape(-1, 1)
    v= np.squeeze(v)  # 去除多余的维度
    print("r1",r1)
    print("r2",r2)
    print("v",v)
    dx = 2 * np.pi / N
    elliptheta = v[4]
    cos_theta = np.cos(elliptheta)
    sin_theta = np.sin(elliptheta)
    Rad = np.array([[cos_theta, sin_theta], [-sin_theta, cos_theta]]).T
    
    ellipX = np.zeros(N)
    ellipY = np.zeros(N)
    for i in range(N):
        ang = i * dx
        x = v[0] * np.cos(ang)
        y = v[1] * np.sin(ang)
        #d11 = Rad @ np.array([[x], [y]])
        #d11 = np.dot(Rad, np.array([[x], [y]]))    
        d11 = np.matmul(Rad, np.array([x, y]))
        
        ellipX[i] = d11[0] + v[2]
        ellipY[i] = d11[1] + v[3]
    print("d11",d11)
    print("d11",d11[0])

    newX = ellipX
    newY = ellipY
    #print("newX",newX)
    # print("newY",newY)
    
    return newX, newY, v