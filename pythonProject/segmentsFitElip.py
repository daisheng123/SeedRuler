# -*- coding: utf-8 -*-
"""
Created on Thu Oct 26 11:42:06 2023

@author: 15959
"""

import numpy as np
from scipy.ndimage import label, binary_erosion
from scipy.ndimage.measurements import label as bwlabel
from skimage.measure import regionprops
from skimage.draw import ellipse_perimeter
import FitEllip
import matplotlib.pyplot as plt
import pandas as pd
import openpyxl
import cv2

def segmentsFitElip(image_part_i, contour_segments, one_grain_area, opt_perimeter):
    len_ = len(contour_segments)
    #print(contour_segments.shape)
    print(len_)
    tmp = np.zeros(len_)
    for kk in range(len_):
        data = contour_segments[kk]

        x = data[:, 0]
        y = data[:, 1]
        num = len(x)
        print("num",num)
        if (np.max(x) - np.min(x)) < 5 or (np.max(y) - np.min(y)) < 5:
            continue
        tmp[kk-1] = num
        k1 = (y[0] - y[round(0.8*num)]) / (x[0] - x[round(0.8*num)])
        k2 = (y[round(num/2)] - y[-1]) / (x[round(num/2)] - x[-1])
        if np.abs(k1 - k2) < 0.3:
            continue
        
        #newX, newY, v = FitEllip.FitEllip(x, y, num) 
        
        # 测试
        # # 打开Excel文件
        # workbook = openpyxl.load_workbook('D:\B_GENERAL\coding transform\project\matlabGUI\data1.xlsx')
        
        # # 选择要读取的工作表
        # worksheet = workbook['Sheet1']
        
        # # 创建一个空数组用于存储数据
        # newX = []
        
        # # 遍历工作表的每一行
        # for row in worksheet.iter_rows(values_only=True):
        #     newX.append(row)
        # newX=np.array(newX).flatten()
        # workbook = openpyxl.load_workbook('D:\B_GENERAL\coding transform\project\matlabGUI\data.xlsx')
        
        # # 选择要读取的工作表
        # worksheet = workbook['Sheet']
        
        # # 创建一个空数组用于存储数据
        # newY = []
        
        # # 遍历工作表的每一行
        # for row in worksheet.iter_rows(values_only=True):
        #     newY.append(row)
        # newY=np.array(newY).flatten()
        
        # data1 = pd.read_excel(r'D:\B_GENERAL\coding transform\project\matlabGUI\data.xlsx')

        # # 将数据转换为数组
        # newY = data1.values
        
        p_thr = opt_perimeter * 0.5

        if num > p_thr:
            # 绘制散点图
            #plt.plot(newX, newY, '.r-', linewidth=1, markersize=2)
            image_part_i=image_part_i.astype(np.uint8)
            #plt.imshow(image_part_i,cmap=plt.cm.gray),plt.title('Pole')
            
            #方式一
            ellipse = cv2.fitEllipse(data)
            cv2.ellipse(image_part_i, ellipse, (0, 255, 0), 2)
            
            #方式二
            #image_part_i[np.round(newY).astype(int), np.round(newX).astype(int)] = 0
            
            #plt.imshow(image_part_i,cmap=plt.cm.gray),plt.title('Process')

            # resized_image = cv2.resize(image_part_i.astype(np.uint8)*255, (0, 0), fx=0.5, fy=0.5)
            # cv2.imshow("xiaoguo",resized_image)
            # cv2.waitKey(0)
    new_image_part_i = image_part_i
    return new_image_part_i