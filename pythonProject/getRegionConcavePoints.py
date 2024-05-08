# -*- coding: utf-8 -*-
"""
Created on Thu Oct 26 11:02:09 2023

@author: 15959
"""

import numpy as np
import cv2
from scipy.ndimage import label, binary_erosion
from scipy.ndimage.measurements import label as bwlabel
from skimage.measure import regionprops
from skimage.measure import label, regionprops
from skimage.morphology import binary_erosion, binary_dilation
import scipy.ndimage as ndimage
from scipy.ndimage import maximum_filter
from scipy.signal import argrelextrema
from scipy.spatial import cKDTree
from scipy.spatial import KDTree
import matplotlib.pyplot as plt
from scipy.interpolate import make_interp_spline
from scipy.interpolate import interp1d
import openpyxl
from scipy.interpolate import UnivariateSpline
from scipy.optimize import curve_fit

def getRegionConcavePoints(image_part_i):
    #l, n = ndimage.label(image_part_i, structure=np.ones((3, 3), dtype=np.int))
    n, l = cv2.connectedComponents(image_part_i.astype(np.uint8), connectivity=8)

    boundaryBw = np.logical_and(image_part_i, np.logical_not(binary_erosion(image_part_i)))
    num, labels, stats, _ = cv2.connectedComponentsWithStats(boundaryBw.astype(np.uint8), connectivity=8)

    # 设置面积阈值
    area_threshold = 5
    
    # 遍历连通区域的统计信息
    for label in range(1, num):
        area = stats[label, cv2.CC_STAT_AREA]
        if area <= area_threshold:
            # 将面积小于阈值的连通区域设置为背景
            labels[labels == label] = 0
    
    # 将背景区域设置为 0，前景区域设置为 255
    boundaryBw=labels.astype(bool)         
    
    b_labeled, b_numObjects = ndimage.label(boundaryBw, structure=np.ones((3, 3), dtype=int))
    #plt.imshow(boundaryBw,cmap='gray')
    #plt.gcf().set_dpi(100)  # 设置图像分辨率
    
    x = []
    y = []

    for k in range(1, b_numObjects+1):
        tmp = (b_labeled == k)
        tmp=tmp.astype(np.uint8)
        # 寻找轮廓
        contours, _ = cv2.findContours(tmp, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)
        
        # 获取第一个连通区域的边界坐标
        bound = contours[0]
        
        # 获取边界坐标的长度
        length = len(bound)
        
        # 获取第一个非孤立像素的索引
        nonzero_indices = np.transpose(np.nonzero(tmp))
        first_nonisolated_index = nonzero_indices[0]
        
        # 调整轮廓点顺序
        bound_start_index = np.argmax(np.all(bound == first_nonisolated_index, axis=2))
        bound = np.roll(bound, -bound_start_index, axis=0)
        
        # 找到x坐标最小的点并将其作为第一个点
        x_min_index = np.argmin(bound[:, 0, 0])
        bound = np.roll(bound, -x_min_index, axis=0)
        
        # 顺时针旋转90度
        bound = np.flip(bound, axis=1)
        bound = np.flip(bound.transpose(1, 0, 2), axis=0)
   
        # 将边界坐标转换为NumPy数组
        b = bound.reshape(-1, 2)
        
        # 提取列坐标和行坐标
        x = b[:, 0] #列坐标
        y = b[:, 1] #行坐标
        x = x[::-1]
        y = y[::-1]
        
# =============================================================================
#         # 计算凸包
#         hull = cv2.convexHull(bound, returnPoints=False)
#         
#         # 计算凹点
#         defects = cv2.convexityDefects(bound, hull)
#         
#         # 获取凹点数量
#         num_defects = defects.shape[0]
#         # 显示凹点
#         for i in range(num_defects):
#             f = defects[i, 0, 2]
#             d = defects[i, 0, 3]
#             #cv2.circle(tmp, start, 3, (0, 0, 255), -1)
#             if(d>1000):
#                 c_row = y[f]  # 当前凹点的行坐标
#                 c_colum = x[f]  # 当前凹点的列坐标
#                 #print('j',j)
#                 print('c_row',c_row)
#                 print('c_colum',c_colum)
#                 #plt.plot(c_colum,c_row , '.', markersize=1)
#                 start = (c_colum,c_row)
#                 cv2.circle(tmp,start, 3, (0, 0, 255), -1)
#         # # 显示结果
#         cv2.imshow("Defects", tmp.astype(np.uint8)*255)
#         cv2.waitKey(0)
#         cv2.destroyAllWindows()
# =============================================================================
        # 测试
        # workbook = openpyxl.load_workbook('D:\B_GENERAL\coding transform\project\matlabGUI\data1.xlsx')
       
        # # 选择要读取的工作表
        # worksheet = workbook['Sheet1']
        
        # # 创建一个空数组用于存储数据
        # x = []
        
        # # 遍历工作表的每一行
        # for row in worksheet.iter_rows(values_only=True):
        #     x.append(row)
        # x=np.array(x).flatten()
        
        # workbook = openpyxl.load_workbook('D:\B_GENERAL\coding transform\project\matlabGUI\data3.xlsx')
       
        # # 选择要读取的工作表
        # worksheet = workbook['Sheet1']
        
        # # 创建一个空数组用于存储数据
        # y = []
        
        # # 遍历工作表的每一行
        # for row in worksheet.iter_rows(values_only=True):
        #     y.append(row)
        # y=np.array(y).flatten()

        angle_threshold = 220  # 角度阈值
        def is_white_pixel(point, image):
            x, y = point
            pixel_value = image[y, x]
            if pixel_value > 0:  # 如果像素值为255，则表示为白色
                return True
            else:
                return False

        step = 7
        # 转置列坐标和行坐标
        x1 = x.reshape(-1, 1)
        y1 = y.reshape(-1, 1)
        # 将 x1 转换为一维数组
        x1 = x1.flatten()
        y1 = y1.flatten()
        len_ = len(y1) + 2*step
        new_x = np.zeros(len_)
        new_x[step:len_-step] = x1
        new_x[0:step] = x1[-step:]
        new_x[len_-step:len_] = x1[0:step]
        new_y = np.zeros(len_)
        new_y[step:len_-step] = y1
        new_y[0:step] = y1[-step:]
        new_y[len_-step:len_] = y1[0:step]
        arc_length = np.zeros(len(y1))
        for j in range(len(y1)):
            
            radius = 8.6

            # # 方法二
            # def line_func(x, a, b):
            #     return a * x + b
            #
            # def calculate_intersection(p , xdata, ydata,flag):
            #     r = radius
            #     # Fit a line to the points p, p_prev, p_next
            #     params, _ = curve_fit(line_func, xdata, ydata)
            #     a, b = params
            #
            #     if (np.max(xdata) - np.min(xdata) <= 2):
            #         a = 20000
            #     if (np.max(xdata) - np.min(xdata) <= 2)and(flag):
            #         a = 18000
            #
            #     # Calculate the intersection of the line and circle C
            #     x_center = new_x[j+step]
            #     y_center = new_y[j+step]
            #     x_intersects = np.roots(
            #         [1 + a ** 2, -2 * x_center - 2 * a * (b - y_center), x_center ** 2 + (b - y_center) ** 2 - r ** 2])
            #     y_intersects = line_func(x_intersects, a, b)
            #
            #     # Find the intersection point closest to p_prev
            #     dist_prev = np.sqrt((p[0] - x_intersects) ** 2 + (p[1] - y_intersects) ** 2)
            #     closest_idx = np.argmin(dist_prev)
            #     intersection_point = [x_intersects[closest_idx], y_intersects[closest_idx]]
            #
            #     return intersection_point
            #
            # p=[new_x[j+step], new_y[j+step]]
            # intersection_point_1 = calculate_intersection(p, new_x[j:j+step+1], new_y[j:j+step+1],False)
            # p=[new_x[j+2*step], new_y[j+2*step]]
            # intersection_point_2 = calculate_intersection(p, new_x[j+step:j+2*step+1], new_y[j+step:j+2*step+1],True)
            #
            # A_Point_x = intersection_point_1[0]
            # A_Point_y = intersection_point_1[1]
            # B_Point_x = intersection_point_2[0]
            # B_Point_y = intersection_point_2[1]

            # 方法一
            later_vec = np.vstack((new_x[j:j+step+1], new_y[j:j+step+1]))  # 第一行为x坐标，第二行为y坐标
            pre_vec = np.vstack((new_x[j+step:j+2*step+1], new_y[j+step:j+2*step+1]))
            later_unique_x, later_unique_idx = np.unique(later_vec[0], return_index=True)
            later_unique_y = later_vec[1][later_unique_idx]
            pre_unique_x, pre_unique_idx = np.unique(pre_vec[0], return_index=True)
            pre_unique_y = pre_vec[1][pre_unique_idx]


            later_interp_func = interp1d(later_unique_x, later_unique_y, kind='linear')
            later_fited_y = later_interp_func(later_vec[0])

            pre_interp_func = interp1d(pre_unique_x, pre_unique_y, kind='linear')
            pre_fited_y = pre_interp_func(pre_vec[0])

            later_coeff = np.polyfit(later_vec[0], later_fited_y, deg=1)
            pre_coeff = np.polyfit(pre_vec[0], pre_fited_y, deg=1)

            a1 = later_coeff[0]
            a2 = pre_coeff[0]

            if np.max(later_vec[0, :]) - np.min(later_vec[0, :]) <= 2:
                a1 = 20000
            if np.max(pre_vec[0, :]) - np.min(pre_vec[0, :]) <= 2:
                a2 = 18000

            l_vec = np.array([later_vec[0, -4]-later_vec[0, -1], later_vec[1, -4]-later_vec[1, -1], 0])
            p_vec = np.array([pre_vec[0, -4]-pre_vec[0, 0], pre_vec[1, -4]-pre_vec[1, 0], 0])
            B_Point_x1 = -np.sqrt((radius)**2/(1+a1**2))
            B_Point_y1 = a1 * B_Point_x1
            B_Point_x2 = np.sqrt((radius)**2/(1+a1**2))
            B_Point_y2 = a1 * B_Point_x2
            d1 = (B_Point_x1 - l_vec[0])**2 + (B_Point_y1 - l_vec[1])**2
            d2 = (B_Point_x2 - l_vec[0])**2 + (B_Point_y2 - l_vec[1])**2
            if d1 < d2:
                B_Point_x = B_Point_x1
                B_Point_y = B_Point_y1
            else:
                B_Point_x = B_Point_x2
                B_Point_y = B_Point_y2
            A_Point_x1 = -np.sqrt((radius)**2/(1+a2**2))
            A_Point_y1 = a2 * A_Point_x1
            A_Point_x2 = np.sqrt((radius)**2/(1+a2**2))
            A_Point_y2 = a2 * A_Point_x2
            d3 = (A_Point_x1 - p_vec[0])**2 + (A_Point_y1 - p_vec[1])**2
            d4 = (A_Point_x2 - p_vec[0])**2 + (A_Point_y2 - p_vec[1])**2
            if d3 < d4:
                A_Point_x = A_Point_x1
                A_Point_y = A_Point_y1
            else:
                A_Point_x = A_Point_x2
                A_Point_y = A_Point_y2

            # # 方法一
            # angles = []
            # for theta in range(1, 361):
            #     m = radius * np.cos(np.radians(theta))
            #     n = radius * np.sin(np.radians(theta))
            #     point = [int(m + x[j]), int(n + y[j])]
            #     angles.append(point)
            # distances1 = [np.sqrt((A_Point_x - point[0]) ** 2 + (A_Point_y - point[1]) ** 2) for point in angles]
            # distances2 = [np.sqrt((B_Point_x - point[0]) ** 2 + (B_Point_y - point[1]) ** 2) for point in angles]
            # closest_point1 = angles[np.argmin(distances1)]
            # closest_point2 = angles[np.argmin(distances2)]
            #
            # # A_index = angles.index(closest_point1)
            # # B_index = angles.index(closest_point2)
            #
            # arc1 = angles[angles.index(closest_point1):angles.index(closest_point2) + 1]
            # arc2 = angles[angles.index(closest_point2):] + angles[:angles.index(closest_point1) + 1]
            # n1 = np.sum([1 for point in arc1 if is_white_pixel(point, l)])
            # n2 = np.sum([1 for point in arc2 if is_white_pixel(point, l)])
            # fore_arc_length = max(n1, n2)
            # arc_length[j] = fore_arc_length

            # 方法二
            theta = np.arange(1, 361)
            # # 假设 radius 和 theta 是已知的变量
            #
            # 计算模板圆上点的 x 坐标
            circle_x = radius * np.cos(np.deg2rad(theta))
            # 计算模板圆上点的 y 坐标
            circle_y = -radius * np.sin(np.deg2rad(theta))
            # 将 x 和 y 坐标合并为一个数组
            circle = np.column_stack((circle_x, circle_y))
            #print("circle",circle)
            # 创建一个临时数组
            tmp = np.array([[A_Point_x, A_Point_y], [B_Point_x, B_Point_y]])
            #print("tmp",tmp)
            # 使用 dsearchn 函数计算最近邻点的索引和距离

            tree = cKDTree(circle)
            dist, index = tree.query(tmp, k=1)
            index = index.astype(int)
            r_x = circle[index, 0]
            r_y = circle[index, 1]
            A_x = r_x[0]
            A_y = r_y[0]
            B_x = r_x[1]
            B_y = r_y[1]
            A_index = index[0]
            B_index = index[1]

            if A_index < B_index:
                fore_arc_length = A_index + (360 - B_index)
                length1 = B_index - A_index
                length2 = 360 - length1
                n2 = 0
                n3 = 0
                for ind in range(A_index, B_index+1):
                    if l[int(round(circle_y[ind]+later_vec[1,-1])), int(round(circle_x[ind]+later_vec[0,-1]))] > 0:
                        n2 += 1

                for ind in range(0,A_index+1):
                    if l[int(round(circle_y[ind]+later_vec[1,-1])), int(round(circle_x[ind]+later_vec[0,-1]))] > 0:
                        n3 += 1

                for ind in range(B_index, 360):
                    if l[int(round(circle_y[ind]+later_vec[1,-1])), int(round(circle_x[ind]+later_vec[0,-1]))] > 0:
                        n3 += 1

                if n2 < n3:
                    fore_arc_length = n3
                else:
                    fore_arc_length = n2

            else:
                fore_arc_length = A_index - B_index
                n2 = 0
                for ind in range(B_index, A_index+1):
                    if l[int(round(circle_y[ind]+later_vec[1,-1])), int(round(circle_x[ind]+later_vec[0,-1]))] > 0:
                        n2 += 1

                n3 = 0
                for ind in range(0,B_index+1):
                    if l[int(round(circle_y[ind]+later_vec[1,-1])), int(round(circle_x[ind]+later_vec[0,-1]))] > 0:
                        n3 += 1

                for ind in range(A_index, 360):
                    if l[int(round(circle_y[ind]+later_vec[1,-1])), int(round(circle_x[ind]+later_vec[0,-1]))] > 0:
                        n3 += 1

                if n2 < n3:
                    fore_arc_length = n3
                else:
                    fore_arc_length = n2

            if A_index == B_index:
                fore_arc_length = 350

            arc_length[j] = fore_arc_length
            
        
        print("hello")
        #print("later_vec",later_vec)
        #print("pre_vec",pre_vec)
        # print("later_fited_y",later_fited_y)
        #print("a1",a1)
        # print("pre_fited_y",pre_fited_y)
        #print("a2",a2)
        # print("x1",x1)
        # print("y1",y1)
        # print("x1_",len(x1))
        # #print("new_x",new_x)
        # #print("new_y",new_y)
        #print("index",index)
        # print("A_index",A_index)
        # print("B_index",B_index)
        # print("n2",n2)
        # print("n3",n3)
        #print("fore_arc_length",fore_arc_length)
        #print("arc_length",arc_length)
        

        arc_length_abs = np.abs(arc_length)
                        
        local_maxima = argrelextrema(arc_length_abs, np.greater, order=8)[0]
        TF = np.zeros_like(arc_length, dtype=bool)
        TF[local_maxima] = True
        
        num_1 = np.sum(TF)
        print("TF",num_1)
        
        for t in range(len(y1)):
            if arc_length[t] < 220:
                TF[t] = False

        len_ = len(y1)
        num_2 = np.sum(TF)
        print("TF",num_2)
    return TF,len_,x,y
